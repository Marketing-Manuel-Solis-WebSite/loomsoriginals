import io
import sys
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

path = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\Carlos Anaya Ruiz\Downloads\Untitled spreadsheet (11).xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n═════════ SHEET: {sheet_name}  ({ws.max_row} rows × {ws.max_column} cols) ═════════")
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Print up to 500 chars per row
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            else:
                s = str(v)
                if len(s) > 160:
                    s = s[:160] + "…"
                cells.append(s)
        line = " | ".join(cells)
        if len(line) > 600:
            line = line[:600] + "…"
        print(f"{row_idx:>3}: {line}")
        if row_idx > 200:
            print("… (truncated at row 200)")
            break
