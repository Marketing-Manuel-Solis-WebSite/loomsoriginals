"""Read the full catalog and classify each video into a series + category."""
import io
import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path(r"C:\Users\Carlos Anaya Ruiz\Downloads\Untitled spreadsheet (11).xlsx")

# ─────────────────────────────────────────────────────────────────────
# Videos that are already in the DB (Uniendo Familias EP 1-3) — skip.
# ─────────────────────────────────────────────────────────────────────
ALREADY_IN_DB = {"PmU1yOfB9C8", "3Z6BOOCBgas", "n8aJSPHdTXg"}

# Skip internal / private videos.
SKIP_PATTERNS = [
    re.compile(r"looms", re.I),
    re.compile(r"configurar.*firma", re.I),
    re.compile(r"guía.*outlook", re.I),
    re.compile(r"guia.*outlook", re.I),
]

# Uniendo Familias extras (teasers, behind the scenes).
UF_TEASER_PATTERN = re.compile(
    r"(uniendo familias|teaser.*familias|familias.*teaser|reencuentro.*30|abrazo.*esperar)",
    re.I,
)

# Patterns used for classification.
CLASS_RULES = [
    # (series_slug, category_slugs[], regex)
    (
        "testimonios-reales",
        ["casos-reales", "reunificacion-familiar"],
        re.compile(
            r"(testimonio|caso real|residencia (permanente|aprobada)|aprobaron mi residencia|"
            r"green.?card|greencard|experiencia al obtener|la historia de |"
            r"kimberly flores|yaztiri morales|rocío dávila|rocio davila|ramiro escobedo|"
            r"rosa arredondo|maría lourdes|maria lourdes|agustín quiroz|agustin quiroz|"
            r"jesús|jose alvarado|josé alvarado|hoy (soy|ya soy) residente|"
            r"sentí mucha alegría|senti mucha alegria|visa de matrimonio|pensó que ese momento|"
            r"años lejos de su familia|24 años lejos|después de años|abrazo que esperar|"
            r"se borran fácil|reencuentro|hoy me aprobaron)",
            re.I,
        ),
    ),
    (
        "guia-migratoria",
        [],  # categories per-video
        re.compile(
            r"(visa t|visa u|vawa|visa humanitaria|sijs|visa juvenil|formulario g-?28|"
            r"seguro social falso|i-275|i-130|ice detuvo|detenido por ice|corte de inmigración|"
            r"corte de inmigracion|estatus migratorio|errores migratorios|notice to appear|"
            r"habeas ?corpus|negligencia médica|negligencia medica|visa e-?2|tormenta|"
            r"accidentes (automovi|de auto)|explosi[oó]n en texas|daños por tormenta|"
            r"danos por tormenta|¿sab[ií]as que|¿qu[eé] pasa si|¿puedo|abogado civil|"
            r"juez federal|acompañante militar|familiar militar|mitos|infórmate|"
            r"mito|verdad o mito|inmigrante juvenil|boletín|cambios en la visa|"
            r"tú llamada|llamada está|habeas corpus|abuso financiero|emocional.*vawa|"
            r"reflexionar|invertir y emprender|emprender en ee.?uu|explotación laboral|"
            r"explotacion laboral|sin depender|dramatización|dramatizacion)",
            re.I,
        ),
    ),
    (
        "en-vivo",
        [],
        re.compile(
            r"(\blive\b|en vivo|podcast|despierta houston|consejos con|live podcast|"
            r"boletín informativo|boletin informativo|miércoles|miercoles.*1:00pm|"
            r"abogado responde y aclara|live\s*podcast)",
            re.I,
        ),
    ),
    (
        "el-bufete",
        [],
        re.compile(
            r"(35 años de experiencia|manuel solís: tu equipo|manuel solis: tu equipo|"
            r"oficina del abogado|bufete de abogados|por qué recomiendan|"
            r"por que recomiendan|confía en el bufete|confia en el bufete|"
            r"tranquilidad migratoria|equipo legal)",
            re.I,
        ),
    ),
]

# Sub-categorization for Guía Migratoria videos.
SUB_CAT_RULES = [
    (re.compile(r"visa t|explotación laboral|explotacion laboral", re.I), ["asilo"]),
    (re.compile(r"visa u|víctim|victim", re.I), ["asilo"]),
    (re.compile(r"vawa|violencia", re.I), ["asilo"]),
    (re.compile(r"sijs|juvenil", re.I), ["asilo"]),
    (re.compile(r"ice|habeas|detenido|detención|detencion|notice to appear|corte de inmigración|corte de inmigracion", re.I), ["deportacion"]),
    (re.compile(r"i-130|reunif|familiar militar|acompañante militar", re.I), ["reunificacion-familiar"]),
    (re.compile(r"visa e-?2|invertir|emprender|visa e2", re.I), ["visas-de-trabajo"]),
    (re.compile(r"ciudadan|naturalización|naturalizacion", re.I), ["ciudadania"]),
]


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def slugify(s: str, max_len: int = 80) -> str:
    s = strip_accents(s or "").lower()
    s = re.sub(r"[^a-z0-9\s-]", " ", s)
    s = re.sub(r"\s+", "-", s.strip())
    s = re.sub(r"-+", "-", s)
    return s[:max_len].rstrip("-") or "video"


def parse_date(v) -> str | None:
    if not v:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%b %d, %Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).isoformat()
        except ValueError:
            continue
    return None


def classify(title: str) -> tuple[str, list[str]]:
    """Return (series_slug, [category_slugs])."""
    for series_slug, base_cats, rx in CLASS_RULES:
        if rx.search(title):
            cats = list(base_cats)
            if series_slug == "guia-migratoria":
                for sub_rx, sub_cats in SUB_CAT_RULES:
                    if sub_rx.search(title):
                        cats.extend(sub_cats)
                        break
                if not cats:
                    cats = ["casos-reales"]
            if not cats and series_slug == "testimonios-reales":
                cats = ["casos-reales", "reunificacion-familiar"]
            if not cats and series_slug == "en-vivo":
                cats = ["casos-reales"]
            if not cats and series_slug == "el-bufete":
                cats = ["casos-reales"]
            return series_slug, cats
    # fallback bucket
    return "historias-cortas", ["casos-reales"]


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    print(f"Header: {header}")

    manifest = {
        "series": {
            "uniendo-familias-manuel-solis": {
                "title_es": "Uniendo Familias con Manuel Solís",
                "title_en": "Reuniting Families with Manuel Solís",
                "synopsis_es": "Serie insignia de Loom Originals. Historias reales de familias que reconstruyeron su camino migratorio con el acompañamiento del Bufete Manuel Solís.",
                "is_featured": True,
                "featured_order": 1,
            },
            "testimonios-reales": {
                "title_es": "Testimonios Reales",
                "title_en": "Real Testimonies",
                "synopsis_es": "Clientes del Bufete Manuel Solís cuentan en primera persona cómo se transformaron sus vidas al obtener su residencia permanente, reunirse con sus familias o proteger sus derechos en Estados Unidos.",
                "is_featured": False,
                "featured_order": 2,
            },
            "guia-migratoria": {
                "title_es": "Guía Migratoria",
                "title_en": "Immigration Guide",
                "synopsis_es": "Cápsulas informativas sobre procesos migratorios — Visa T, Visa U, VAWA, SIJS, I-130, habeas corpus, defensa contra ICE y más. Presentadas por el equipo legal de Manuel Solís.",
                "is_featured": False,
                "featured_order": 3,
            },
            "en-vivo": {
                "title_es": "En Vivo y Podcasts",
                "title_en": "Live & Podcasts",
                "synopsis_es": "Sesiones en vivo, podcasts y entrevistas del Bufete Manuel Solís donde se responden dudas reales de la comunidad migrante.",
                "is_featured": False,
                "featured_order": 4,
            },
            "el-bufete": {
                "title_es": "El Bufete",
                "title_en": "About the Firm",
                "synopsis_es": "Conozca al equipo, la historia y los valores del Bufete Manuel Solís — 35 años acompañando a familias migrantes.",
                "is_featured": False,
                "featured_order": 5,
            },
            "historias-cortas": {
                "title_es": "Historias Cortas",
                "title_en": "Short Stories",
                "synopsis_es": "Momentos breves — frases, reflexiones y fragmentos del trabajo diario del Bufete Manuel Solís.",
                "is_featured": False,
                "featured_order": 6,
            },
        },
        "videos": [],
    }

    seen_slugs_per_series: dict[str, set[str]] = {}
    ep_counters: dict[str, int] = {}

    for row in rows[2:]:  # skip header + total row
        if not row:
            continue
        yt_id = row[0]
        title = (row[1] or "").strip() if len(row) > 1 else ""
        published = row[2] if len(row) > 2 else None
        duration = row[3] if len(row) > 3 else None

        if not yt_id or not title:
            continue
        yt_id = str(yt_id).strip()
        if yt_id in ALREADY_IN_DB:
            continue
        # skip obvious YouTube IDs that are not 11 chars
        if len(yt_id) != 11:
            continue
        if any(rx.search(title) for rx in SKIP_PATTERNS):
            continue

        series_slug, categories = classify(title)

        # UF teasers are not full episodes — they go into Testimonios Reales.
        if UF_TEASER_PATTERN.search(title) and yt_id not in ALREADY_IN_DB:
            series_slug = "testimonios-reales"
            categories = ["reunificacion-familiar", "casos-reales"]

        # Generate unique slug within series
        base_slug = slugify(title)
        seen = seen_slugs_per_series.setdefault(series_slug, set())
        slug = base_slug
        n = 2
        while slug in seen:
            slug = f"{base_slug}-{n}"
            n += 1
        seen.add(slug)

        ep_counters[series_slug] = ep_counters.get(series_slug, 0) + 1

        manifest["videos"].append({
            "youtube_id": yt_id,
            "series_slug": series_slug,
            "slug": slug,
            "title": title[:200],
            "duration_seconds": int(duration) if duration and str(duration).replace(".", "").isdigit() else None,
            "published_at": parse_date(published),
            "episode_number": ep_counters[series_slug],
            "categories": categories,
        })

    out = Path(__file__).parent / "catalog-manifest.json"
    out.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    # Stats
    by_series: dict[str, int] = {}
    for v in manifest["videos"]:
        by_series[v["series_slug"]] = by_series.get(v["series_slug"], 0) + 1
    print("\n=== Totals ===")
    for k, v in sorted(by_series.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v} videos")
    print(f"\n→ wrote {out} ({len(manifest['videos'])} videos)")


if __name__ == "__main__":
    main()
